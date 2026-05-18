# local_excel_manager.py
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

EXCEL_FILE_PATH = "Job_Search_Tracker.xlsx"

def deep_sanitize_for_xml(val) -> str:
    """Cleans up string objects to make them 100% compliant with Excel XML structures."""
    if val is None:
        return "N/A"
    clean_str = str(val).strip()
    clean_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', clean_str)
    return clean_str

def parse_score_weight(score_display: str) -> int:
    """Extracts the numerator from the score string for numeric descending sorting."""
    try:
        return int(str(score_display).split('/')[0])
    except:
        return 0

def parse_salary_weight(salary_display: str) -> float:
    """Converts Indian denomination text scales (CR/Lakhs) into a normalized numeric value."""
    try:
        text = str(salary_display).upper()
        # Find all decimal or whole numbers
        numbers = [float(n) for n in re.findall(r'\d+\.\d+|\d+', text)]
        if not numbers:
            return 0.0
        
        # Use the upper ceiling bound of the salary bracket for sorting logic
        target_val = max(numbers)
        
        if "CR" in text:
            return target_val * 10_000_000
        if "L" in text or "LAKH" in text:
            return target_val * 100_000
        return target_val
    except:
        return 0.0

def append_to_local_sheet(new_job_details: dict, linkedin_connections: list) -> None:
    """
    Ingests, deduplicates, multi-key sorts, and updates the local job hunting tracker
    without breaking Excel structural layout rules.
    """
    headers = [
        "Company", "Job Title", "Match Score", 
        "Est. Total Comp", "Comp Breakdown", "Job URL", 
        "Target Outreach Links", "Posted Date", "Location", 
        "Workplace Policy", "Match Justification"
    ]
    
    # 1. Sanitize and prepare incoming data row components
    new_company = deep_sanitize_for_xml(new_job_details.get("company", "Unknown"))
    new_title = deep_sanitize_for_xml(new_job_details.get("job_title", "Unknown"))
    
    raw_job_link = str(new_job_details.get("job_url", "")).strip().replace('"', '')
    if not raw_job_link or raw_job_link.upper() == "N/A":
        query_encoded = f"{new_company} {new_title}".replace(" ", "+")
        raw_job_link = f"https://www.google.com/search?q={query_encoded}+jobs"
        
    raw_outreach_link = str(linkedin_connections[0]).strip().replace('"', '') if linkedin_connections else ""

    fresh_job_object = {
        "company": new_company,
        "job_title": new_title,
        "score_display": f"{new_job_details.get('score_out_of_10', 0)}/10",
        "estimated_total_comp": deep_sanitize_for_xml(new_job_details.get("estimated_total_comp", "₹1.25 CR+ (Inferred)")),
        "comp_breakdown": deep_sanitize_for_xml(new_job_details.get("comp_breakdown", "Not Specified")),
        "raw_job_link": raw_job_link,
        "raw_outreach_link": raw_outreach_link,
        "posted_date": deep_sanitize_for_xml(new_job_details.get("posted_date", "2026-05-17")),
        "location": deep_sanitize_for_xml(new_job_details.get("location", "Unknown")),
        "workplace_policy": deep_sanitize_for_xml(new_job_details.get("workplace_policy", "Not Specified")),
        "justification": deep_sanitize_for_xml(new_job_details.get("justification", "N/A"))
    }

    all_jobs = []
    
    # 2. Ingest existing entries from disk if file is initialized
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue
                
            # Duplicate check match criteria
            if (str(row[0].value).lower().strip() == fresh_job_object["company"].lower() and 
                str(row[1].value).lower().strip() == fresh_job_object["job_title"].lower()):
                print(f"Skip duplicate: {fresh_job_object['job_title']} at {fresh_job_object['company']}")
                wb.close()
                return
                
            all_jobs.append({
                "company": str(row[0].value),
                "job_title": str(row[1].value),
                "score_display": str(row[2].value),
                "estimated_total_comp": str(row[3].value),
                "comp_breakdown": str(row[4].value),
                "raw_job_link": row[5].hyperlink.target if row[5].hyperlink else "",
                "raw_outreach_link": row[6].hyperlink.target if row[6].hyperlink else "",
                "posted_date": str(row[7].value),
                "location": str(row[8].value),
                "workplace_policy": str(row[9].value),
                "justification": str(row[10].value)
            })
        wb.close()
    
    # Add the current scrape target record to the array
    all_jobs.append(fresh_job_object)
    
    # 3. Dynamic Multi-Key Sorting Engine execution
    # Level 1: Date Posted -> Ascending (Earliest dates group first)
    # Level 2: Match Score -> Descending (Negative inversion makes highest values rank first)
    # Level 3: Salary Value -> Descending (Negative inversion forces top compensation metrics to lead)
    all_jobs.sort(key=lambda x: (
        x["posted_date"],
        -parse_score_weight(x["score_display"]),
        -parse_salary_weight(x["estimated_total_comp"])
    ))

    # 4. Rewrite the ledger tracking worksheet completely
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Hunting Tracker"
    ws.append(headers)
    
    for job in all_jobs:
        row_data = [
            job["company"],
            job["job_title"],
            job["score_display"],
            job["estimated_total_comp"],
            job["comp_breakdown"],
            "Open Listing ↗" if job["raw_job_link"] else "N/A",
            "Find Recruiters (India) 👥" if job["raw_outreach_link"] else "N/A",
            job["posted_date"],
            job["location"],
            job["workplace_policy"],
            job["justification"]
        ]
        ws.append(row_data)
        current_row = ws.max_row
        
        # Rebind native Hyperlink relationships safely
        if job["raw_job_link"]:
            cell = ws.cell(row=current_row, column=6)
            cell.hyperlink = job["raw_job_link"]
            cell.font = Font(color="0563C1", underline="single")
            
        if job["raw_outreach_link"]:
            cell = ws.cell(row=current_row, column=7)
            cell.hyperlink = job["raw_outreach_link"]
            cell.font = Font(color="0563C1", underline="single")
            
        # Structure display alignment wrapper
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    # Set cell layout dimensions
    column_widths = {
        "A": 20, "B": 30, "C": 12, "D": 22, "E": 40, "F": 18, "G": 28, "H": 15, "I": 25, "J": 20, "K": 60
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(EXCEL_FILE_PATH)
    wb.close()
